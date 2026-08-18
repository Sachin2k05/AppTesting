"""Generate the VoiceHire backend security deliverables.

One source of truth for findings, endpoints and dependencies; emits both the
Markdown reports and the two Excel workbooks so they can never drift apart.

Every finding below was produced by reading the code and, where marked
CONFIRMED-LIVE, by probing a locally running instance read-only. Nothing was
modified, and no credential value is reproduced anywhere in the output.
"""

import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "Vulnerability Test Results")
os.makedirs(OUT, exist_ok=True)

TARGET = "voicehire-backend (Node.js 18 / Express 4 / Supabase Postgres)"
ASSESSED = date.today().isoformat()

# ── Findings ─────────────────────────────────────────────────────────────
# id, severity, type, file, endpoint, description, exploit, impact, fix, status
F = [
    ("VH-001", "Critical", "Security Misconfiguration (CORS)",
     "voicehire-backend/server.js:15-21", "ALL endpoints",
     "The CORS origin callback is `callback(null, true)` for every origin, and "
     "`credentials: true` is set. The server therefore reflects any Origin and "
     "tells the browser credentials may be sent.",
     "An attacker hosts evil.example.com. A signed-in victim visits it. Script on "
     "that page issues cross-origin XHR to the API; the browser permits it because "
     "the server reflects the attacker's origin and allows credentials. Any request "
     "the victim's browser can authenticate, the attacker can read.",
     "Cross-origin read of every authenticated endpoint: profile, knowledge graph, "
     "conversation history, applications. Full account data disclosure.",
     "Replace the permissive callback with an explicit allow-list of the deployed "
     "frontend origins. If credentials are genuinely needed, never combine them with "
     "a reflected origin. Example: `origin: ['https://voicehire.app']`.",
     "CONFIRMED-LIVE"),

    ("VH-002", "High", "Broken Access Control (A01)",
     "voicehire-backend/src/routes/tracker.js, src/routes/notifications.js",
     "/api/tracker/* (8), /api/notifications/* (8)",
     "Sixteen endpoints are mounted with no authMiddleware and read/write "
     "module-level `new Map()` stores (savedJobsStore, remindersStore, notifStore) "
     "that are shared by every caller of the process. There is no per-user scoping "
     "of any kind.",
     "An unauthenticated client calls GET /api/tracker/dashboard or "
     "GET /api/notifications and receives whatever is in the shared store. Write "
     "endpoints (POST /reminder, POST /read, POST /snooze, DELETE /:id, "
     "DELETE /saved/:id) let the same anonymous caller mutate or delete entries "
     "belonging to other users.",
     "Cross-tenant read and write of saved jobs, reminders and notifications with "
     "no credential at all.",
     "Add authMiddleware to every route in both files and key each store by "
     "req.userId, or move them to RLS-protected tables via req.supabase. An "
     "in-memory Map is also lost on restart and not shared across instances.",
     "CONFIRMED-LIVE"),

    ("VH-003", "High", "Unauthenticated Mail Relay / Abuse",
     "voicehire-backend/src/routes/applications.js:277",
     "POST /api/applications/send-email",
     "The endpoint sends email to a caller-supplied `to` address with no "
     "authentication. Only a 20/minute rate limit guards it, and that limit is "
     "bypassable (see VH-004).",
     "An attacker POSTs arbitrary recipients and job/company strings, rotating the "
     "X-Forwarded-For header to defeat throttling, and uses the service to deliver "
     "mail from the project's verified sender.",
     "Spam and phishing sent from the organisation's domain and sender reputation; "
     "possible blocklisting of the sending domain; cost abuse.",
     "Require authMiddleware, and derive the recipient from the authenticated "
     "user's own application record rather than the request body. Never accept a "
     "free-form recipient from an anonymous caller.",
     "CONFIRMED-CODE"),

    ("VH-004", "High", "Rate Limiting Bypass",
     "voicehire-backend/src/middleware/rateLimit.js:26-30 (+ server.js)",
     "All rate-limited endpoints",
     "`_clientKey()` reads `req.headers['x-forwarded-for']` FIRST and only falls "
     "back to req.ip. Express `trust proxy` is never enabled, so that header is "
     "entirely client-controlled and unvalidated.",
     "The attacker sends a different X-Forwarded-For value on each request. Every "
     "request lands in a fresh bucket, so the limiter never triggers.",
     "Throttling is ineffective for unauthenticated callers, enabling brute force, "
     "mail-relay abuse (VH-003) and CPU exhaustion (VH-006).",
     "Enable `app.set('trust proxy', 1)` and use `req.ip` only, letting Express "
     "parse the proxy chain. Do not read the raw header. For multi-instance "
     "deployments move the buckets to a shared store.",
     "CONFIRMED-CODE"),

    ("VH-005", "High", "Vulnerable Dependency (CVE)",
     "voicehire-backend/package.json", "POST /api/applications/send-email",
     "nodemailer is flagged by npm audit as HIGH: 'Email to an unintended domain "
     "can occur due to Interpretation Conflict'. The project exposes an "
     "unauthenticated endpoint that calls it.",
     "A crafted recipient string is parsed differently by the library than by the "
     "application's validation, so mail is delivered to a domain the application "
     "did not intend.",
     "Mail misdelivery, which combined with VH-003 lets an anonymous caller direct "
     "delivery to unintended domains.",
     "Run `npm audit fix`, upgrade nodemailer to the patched release, and add "
     "server-side recipient validation that is independent of the library's parser.",
     "CONFIRMED-SCAN"),

    ("VH-006", "Medium", "Resource Exhaustion / DoS",
     "voicehire-backend/src/routes/applications.js:371",
     "POST /api/applications/generate-resume",
     "PDF generation via PDFKit runs for any anonymous caller with a "
     "caller-supplied profile object. Rate limited to 20/min, but that limit is "
     "bypassable per VH-004.",
     "The attacker submits large profile payloads (up to the 1 MB body limit) in a "
     "loop with rotating X-Forwarded-For headers, forcing continuous CPU-bound PDF "
     "rendering on a single-instance host.",
     "CPU exhaustion and denial of service for legitimate users; on Render's free "
     "tier this alone can make the service unresponsive.",
     "Require authentication, cap the payload well below 1 MB for this route, and "
     "fix the limiter key (VH-004). Consider queueing generation.",
     "CONFIRMED-CODE"),

    ("VH-007", "Medium", "Missing Security Headers",
     "voicehire-backend/server.js:31-34", "ALL endpoints",
     "Four headers are set manually (X-Content-Type-Options, X-Frame-Options, "
     "Referrer-Policy, X-Permitted-Cross-Domain-Policies). Absent: "
     "Strict-Transport-Security, Content-Security-Policy, Permissions-Policy.",
     "Without HSTS a first request over plain HTTP can be intercepted and "
     "downgraded. Without a Permissions-Policy the page's microphone permission is "
     "not constrained for embedded content.",
     "Downgrade attacks on first contact; weaker defence in depth for any HTML the "
     "backend serves.",
     "Adopt `helmet` rather than hand-rolled headers, and enable HSTS "
     "(max-age >= 31536000; includeSubDomains) plus a Permissions-Policy that "
     "restricts microphone to self.",
     "CONFIRMED-LIVE"),

    ("VH-008", "Medium", "Insecure Design (Shared Mutable State)",
     "voicehire-backend/src/routes/tracker.js:5-6, notifications.js",
     "/api/tracker/*, /api/notifications/*",
     "State lives in process memory. Render runs multiple instances and recycles "
     "them; the data is neither shared nor durable, and it is seeded with "
     "hardcoded demo values (appliedCount: 2, 'ABC Technologies').",
     "Not an attack so much as a correctness failure that becomes a security "
     "problem: users may be shown another user's data, or fabricated data "
     "presented as their own record.",
     "Users cannot trust the tracker. For an application that reports whether a job "
     "application was really submitted, showing invented counts is a serious "
     "integrity issue.",
     "Move to RLS-protected Supabase tables scoped by user_id, and delete the "
     "hardcoded demo rows.",
     "CONFIRMED-LIVE"),

    ("VH-009", "Medium", "Vulnerable Dependency (CVE)",
     "voicehire-backend/package-lock.json", "n/a",
     "brace-expansion (transitive) is flagged HIGH for denial of service via "
     "exponential-time expansion; uuid is flagged MODERATE for a missing buffer "
     "bounds check in v3/v5/v6 when a buffer is supplied.",
     "brace-expansion is reachable only through build/dev tooling here, so remote "
     "exploitation is unlikely; uuid is used for identifier generation without a "
     "caller-supplied buffer.",
     "Low practical impact in this codebase, but both are known CVEs that a "
     "dependency review will flag and that could become reachable after refactoring.",
     "Run `npm audit fix`. Pin the patched versions and re-run the audit in CI.",
     "CONFIRMED-SCAN"),

    ("VH-010", "Medium", "Authentication Bypass Risk (opt-in)",
     "voicehire-backend/src/middleware/auth.js (ALLOW_LEGACY_AUTH)",
     "ALL authenticated endpoints",
     "When ALLOW_LEGACY_AUTH=true a self-issued HS256 JWT is accepted and "
     "req.supabase is set to the SERVICE-ROLE client, which bypasses RLS entirely "
     "for the whole request.",
     "If the flag is ever enabled in production and JWT_SECRET leaks or is weak, an "
     "attacker mints a token for any userId and operates with RLS disabled.",
     "Complete compromise of every user's data in that configuration.",
     "Keep the flag false in production and assert it at boot; log loudly if set. "
     "Plan removal now that clients use Supabase sign-in. Current default is safe.",
     "REVIEW-ONLY"),

    ("VH-011", "Low", "Information Disclosure (non-production)",
     "voicehire-backend/server.js:109", "ALL endpoints",
     "When NODE_ENV !== 'production' the error handler logs a redacted shape of "
     "the request body. The redaction helper is applied, so values are not "
     "printed, but the field names of failing requests are.",
     "An operator with log access learns which fields were present on a failing "
     "request. Low value on its own.",
     "Minor metadata disclosure limited to non-production environments.",
     "Acceptable as designed. Ensure NODE_ENV=production is set on every deployed "
     "environment; it currently is in render.yaml.",
     "REVIEW-ONLY"),

    ("VH-012", "Low", "Missing Audit Trail",
     "voicehire-backend/server.js", "ALL endpoints",
     "There is no request id, and no audit log of security-relevant events "
     "(sign-in, account linking, application submission, deletion).",
     "After an incident there is no reliable way to reconstruct who did what.",
     "Impedes incident response and forensics.",
     "Add a request-id middleware and structured audit logging for auth and "
     "write operations, excluding payload values.",
     "REVIEW-ONLY"),
]

# ── Controls verified as WORKING (reported honestly, not padded as findings) ──
PASSED = [
    ("Identity is server-verified", "Bearer tokens are verified against Supabase via "
     "serviceClient.auth.getUser(); identity comes from the verified `sub`, never from "
     "client-supplied fields.", "src/middleware/auth.js"),
    ("JWT forgery rejected", "alg=none and forged HS256 tokens were both rejected with "
     "401 against a running instance.", "DAST"),
    ("Two-client RLS model", "Service-role client is reserved for identity linking; request "
     "handlers use an RLS-scoped client bound to the caller's token.", "src/config/supabase.js"),
    ("No SQL injection surface", "All database access goes through the Supabase JS client, "
     "which parameterises. No string-concatenated SQL was found.", "src/routes/*"),
    ("No hardcoded secrets", "No credential-shaped literals found in source. .env is "
     "gitignored; only .env.example is tracked.", "repository scan"),
    ("Errors do not leak internals", "The handler returns a generic 500 and logs only the "
     "failure shape with values redacted.", "server.js:100-118"),
    ("Legacy credential-free auth closed", "/api/auth/register and /login return 410 Gone "
     "unless ALLOW_LEGACY_AUTH is explicitly enabled.", "src/routes/auth.js"),
    ("Body size limited", "express.json and urlencoded are capped at 1 MB.", "server.js:23-24"),
]

# ── Endpoint inventory ───────────────────────────────────────────────────
E = [
    ("/api/auth/session", "POST", "No", "public", "src/routes/auth.js:45", "Exchanges a verified Supabase token for a session"),
    ("/api/auth/me", "GET", "Yes", "user", "src/routes/auth.js:149", ""),
    ("/api/auth/register", "POST", "No", "public", "src/routes/auth.js:220", "410 Gone unless ALLOW_LEGACY_AUTH"),
    ("/api/auth/login", "POST", "No", "public", "src/routes/auth.js:221", "410 Gone unless ALLOW_LEGACY_AUTH"),
    ("/api/auth/logout", "POST", "No", "public", "src/routes/auth.js:226", ""),
    ("/api/profile", "GET", "Yes", "user", "src/routes/profile.js:125", ""),
    ("/api/profile", "POST", "Yes", "user", "src/routes/profile.js:154", ""),
    ("/api/profile", "PATCH", "Yes", "user", "src/routes/profile.js:194", ""),
    ("/api/graph", "GET", "Yes", "user", "src/routes/graph.js:56", "Knowledge graph snapshot"),
    ("/api/graph", "PUT", "Yes", "user", "src/routes/graph.js:96", ""),
    ("/api/graph/backup", "POST", "Yes", "user", "src/routes/graph.js:177", ""),
    ("/api/graph/backups", "GET", "Yes", "user", "src/routes/graph.js:222", ""),
    ("/api/graph/backup/:id", "GET", "Yes", "user", "src/routes/graph.js:249", "Verify ownership of :id"),
    ("/api/graph", "DELETE", "Yes", "user", "src/routes/graph.js:272", ""),
    ("/api/conversation", "GET", "Yes", "user", "src/routes/conversation.js:11", ""),
    ("/api/conversation", "POST", "Yes", "user", "src/routes/conversation.js:43", ""),
    ("/api/conversation/bulk", "POST", "Yes", "user", "src/routes/conversation.js:84", ""),
    ("/api/conversation", "DELETE", "Yes", "user", "src/routes/conversation.js:129", ""),
    ("/api/applications", "POST", "Yes", "user", "src/routes/applications.js:56", ""),
    ("/api/applications/analytics", "GET", "Yes", "user", "src/routes/applications.js:128", ""),
    ("/api/applications", "GET", "Yes", "user", "src/routes/applications.js:174", ""),
    ("/api/applications/:id", "GET", "Yes", "user", "src/routes/applications.js:198", "Verify ownership of :id"),
    ("/api/applications/:id/status", "PATCH", "Yes", "user", "src/routes/applications.js:225", "Verify ownership of :id"),
    ("/api/applications/send-email", "POST", "NO", "public", "src/routes/applications.js:277", "VH-003 mail relay"),
    ("/api/applications/generate-resume", "POST", "NO", "public", "src/routes/applications.js:371", "VH-006 CPU DoS"),
    ("/api/apply", "POST", "Yes", "user", "src/routes/apply.js:247", ""),
    ("/api/jobs/companies", "GET", "No", "public", "src/routes/jobs.js:83", "Reference data"),
    ("/api/jobs", "GET", "No", "public", "src/routes/jobs.js:95", "Reference data"),
    ("/api/jobs/search", "GET", "No", "public", "src/routes/jobs.js:96", "Reference data"),
    ("/api/jobs/ai", "POST", "NO", "public", "src/routes/jobs.js:124", "LLM proxy - cost abuse risk"),
    ("/api/jobs/intent", "POST", "NO", "public", "src/routes/jobs.js:206", "LLM proxy - cost abuse risk"),
    ("/api/jobs/save", "POST", "Yes", "user", "src/routes/jobs.js:251", ""),
    ("/api/jobs/save/:id", "DELETE", "Yes", "user", "src/routes/jobs.js:287", ""),
    ("/api/jobs/saved", "GET", "Yes", "user", "src/routes/jobs.js:315", ""),
    ("/api/matching/analyze", "POST", "NO", "public", "src/routes/matching.js:30", ""),
    ("/api/matching", "GET", "NO", "public", "src/routes/matching.js:65", ""),
    ("/api/matching/recommendations", "GET", "NO", "public", "src/routes/matching.js:86", ""),
    ("/api/matching/:jobId", "GET", "NO", "public", "src/routes/matching.js:107", ""),
    ("/api/tracker/dashboard", "GET", "NO", "public", "src/routes/tracker.js:23", "VH-002"),
    ("/api/tracker/saved", "GET", "NO", "public", "src/routes/tracker.js:52", "VH-002"),
    ("/api/tracker/saved/:id", "DELETE", "NO", "public", "src/routes/tracker.js:65", "VH-002 anonymous delete"),
    ("/api/tracker/interviews", "GET", "NO", "public", "src/routes/tracker.js:79", "VH-002"),
    ("/api/tracker/offers", "GET", "NO", "public", "src/routes/tracker.js:102", "VH-002"),
    ("/api/tracker/rejected", "GET", "NO", "public", "src/routes/tracker.js:113", "VH-002"),
    ("/api/tracker/analytics", "GET", "NO", "public", "src/routes/tracker.js:130", "VH-002"),
    ("/api/tracker/reminder", "POST", "NO", "public", "src/routes/tracker.js:153", "VH-002 anonymous write"),
    ("/api/notifications", "GET", "NO", "public", "src/routes/notifications.js:52", "VH-002"),
    ("/api/notifications/unread", "GET", "NO", "public", "src/routes/notifications.js:69", "VH-002"),
    ("/api/notifications/daily", "GET", "NO", "public", "src/routes/notifications.js:85", "VH-002"),
    ("/api/notifications/weekly", "GET", "NO", "public", "src/routes/notifications.js:102", "VH-002"),
    ("/api/notifications/read", "POST", "NO", "public", "src/routes/notifications.js:120", "VH-002 anonymous write"),
    ("/api/notifications/snooze", "POST", "NO", "public", "src/routes/notifications.js:141", "VH-002 anonymous write"),
    ("/api/notifications/:id", "DELETE", "NO", "public", "src/routes/notifications.js:158", "VH-002 anonymous delete"),
    ("/api/notifications/preferences", "POST", "NO", "public", "src/routes/notifications.js:174", "VH-002 anonymous write"),
]

# ── Dependencies ─────────────────────────────────────────────────────────
D = [
    ("nodemailer", "^6.9.7", "High", "Email to an unintended domain (Interpretation Conflict)",
     "Direct", "Reachable via public POST /api/applications/send-email", "npm audit fix / upgrade"),
    ("brace-expansion", "transitive", "High", "DoS via exponential-time expansion",
     "Transitive", "Build/dev tooling only; not on a request path", "npm audit fix"),
    ("uuid", "^9.0.0", "Moderate", "Missing buffer bounds check in v3/v5/v6 when buf supplied",
     "Direct", "Not called with a caller-supplied buffer", "npm audit fix"),
    ("(1 low advisory)", "-", "Low", "See `npm audit` output for detail",
     "Transitive", "No known request-path exposure", "npm audit fix"),
]

SEV_ORDER = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
counts = {s: sum(1 for f in F if f[1] == s) for s in ["Critical", "High", "Medium", "Low"]}

# Score: start at 100, subtract weighted severity. Documented, not arbitrary.
WEIGHT = {"Critical": 20, "High": 10, "Medium": 4, "Low": 1}
score = 100 - sum(WEIGHT[f[1]] for f in F)
score = max(score, 0)

# ═════════════════════════════════════════════════════════════════════════
# Excel: findings.xlsx  (4 sheets)
# ═════════════════════════════════════════════════════════════════════════
FONT = "Arial"
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
SEV_FILL = {
    "Critical": PatternFill("solid", fgColor="C00000"),
    "High": PatternFill("solid", fgColor="FF7C80"),
    "Medium": PatternFill("solid", fgColor="FFE699"),
    "Low": PatternFill("solid", fgColor="D9E1F2"),
}
SEV_FONT = {"Critical": Font(name=FONT, size=10, bold=True, color="FFFFFF")}


def sheet(wb, title, headers, rows, widths, wrap_cols=(), first=False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font, cell.border = HDR_FILL, HDR_FONT, BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name=FONT, size=10)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top",
                                       wrap_text=(headers[c - 1] in wrap_cols))
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    return ws


wb = Workbook()

# Sheet 1 — Security Findings
fh = ["ID", "Severity", "Vulnerability Type", "File Path", "Endpoint", "Description",
      "Exploitation Scenario", "Impact", "Recommended Fix", "Verification"]
frows = [list(f) for f in sorted(F, key=lambda x: SEV_ORDER[x[1]])]
ws = sheet(wb, "Security Findings", fh, frows,
           [10, 11, 30, 44, 30, 60, 60, 46, 60, 17],
           wrap_cols=("Description", "Exploitation Scenario", "Impact",
                      "Recommended Fix", "File Path", "Endpoint"), first=True)
for r in range(2, len(frows) + 2):
    sev = ws.cell(row=r, column=2).value
    ws.cell(row=r, column=2).fill = SEV_FILL[sev]
    if sev in SEV_FONT:
        ws.cell(row=r, column=2).font = SEV_FONT[sev]

# Sheet 2 — Endpoint Inventory
eh = ["Endpoint", "HTTP Method", "Authentication Required", "Expected Roles",
      "Controller / File Path", "Security Note"]
erows = [list(e) for e in E]
ws2 = sheet(wb, "Endpoint Inventory", eh, erows, [38, 13, 24, 15, 40, 34],
            wrap_cols=("Security Note",))
warn = PatternFill("solid", fgColor="FF7C80")
for r in range(2, len(erows) + 2):
    if ws2.cell(row=r, column=3).value == "NO":
        ws2.cell(row=r, column=3).fill = warn

# Sheet 3 — Dependency Vulnerabilities
dh = ["Package", "Version", "Severity", "Advisory", "Type", "Reachability", "Remediation"]
ws3 = sheet(wb, "Dependency Vulnerabilities", dh, [list(d) for d in D],
            [22, 14, 11, 56, 12, 46, 30], wrap_cols=("Advisory", "Reachability"))
for r in range(2, len(D) + 2):
    sev = ws3.cell(row=r, column=3).value
    if sev in SEV_FILL:
        ws3.cell(row=r, column=3).fill = SEV_FILL[sev]

# Sheet 4 — Risk Summary (formulas over sheet 1)
LAST = len(frows) + 1
SEVR = f"'Security Findings'!$B$2:$B${LAST}"
rs = wb.create_sheet("Risk Summary")
rs["A1"] = "VoiceHire Backend — Risk Summary"
rs["A1"].font = Font(name=FONT, size=14, bold=True, color="1F3864")
rs["A2"] = f"Target: {TARGET}"
rs["A3"] = f"Assessed: {ASSESSED}"
for c in ("A2", "A3"):
    rs[c].font = Font(name=FONT, size=10, italic=True, color="595959")

r = 5
rs.cell(row=r, column=1, value="Severity").font = Font(name=FONT, size=10, bold=True)
rs.cell(row=r, column=2, value="Count").font = Font(name=FONT, size=10, bold=True)
r += 1
for sev in ["Critical", "High", "Medium", "Low"]:
    rs.cell(row=r, column=1, value=sev).font = Font(name=FONT, size=10)
    rs.cell(row=r, column=1).fill = SEV_FILL[sev]
    if sev in SEV_FONT:
        rs.cell(row=r, column=1).font = SEV_FONT[sev]
    rs.cell(row=r, column=2, value=f'=COUNTIF({SEVR},A{r})').font = Font(name=FONT, size=10)
    r += 1
rs.cell(row=r, column=1, value="TOTAL").font = Font(name=FONT, size=10, bold=True)
rs.cell(row=r, column=2, value=f"=SUM(B{r-4}:B{r-1})").font = Font(name=FONT, size=10, bold=True)
total_row = r
r += 2

rs.cell(row=r, column=1, value="Security Score (out of 100)").font = Font(name=FONT, size=10, bold=True)
rs.cell(row=r, column=2,
        value=f"=MAX(0,100-(B{total_row-4}*{WEIGHT['Critical']}+B{total_row-3}*{WEIGHT['High']}"
              f"+B{total_row-2}*{WEIGHT['Medium']}+B{total_row-1}*{WEIGHT['Low']}))"
        ).font = Font(name=FONT, size=12, bold=True)
r += 1
rs.cell(row=r, column=1,
        value=f"Weighting: Critical -{WEIGHT['Critical']}, High -{WEIGHT['High']}, "
              f"Medium -{WEIGHT['Medium']}, Low -{WEIGHT['Low']} from a base of 100."
        ).font = Font(name=FONT, size=9, italic=True, color="595959")
r += 2

rs.cell(row=r, column=1, value="Endpoint exposure").font = Font(name=FONT, size=10, bold=True)
r += 1
for label, formula in [
    ("Total endpoints", f"=COUNTA('Endpoint Inventory'!$A$2:$A${len(E)+1})"),
    ("Requiring authentication", f"=COUNTIF('Endpoint Inventory'!$C$2:$C${len(E)+1},\"Yes\")"),
    ("Unauthenticated (flagged)", f"=COUNTIF('Endpoint Inventory'!$C$2:$C${len(E)+1},\"NO\")"),
    ("Unauthenticated (by design)", f"=COUNTIF('Endpoint Inventory'!$C$2:$C${len(E)+1},\"No\")"),
]:
    rs.cell(row=r, column=1, value=label).font = Font(name=FONT, size=10)
    rs.cell(row=r, column=2, value=formula).font = Font(name=FONT, size=10)
    r += 1
r += 1

rs.cell(row=r, column=1, value="Controls verified as working").font = Font(name=FONT, size=10, bold=True)
r += 1
for name, detail, where in PASSED:
    rs.cell(row=r, column=1, value="PASS  " + name).font = Font(name=FONT, size=10, color="006100")
    rs.cell(row=r, column=2, value=where).font = Font(name=FONT, size=9, color="595959")
    r += 1

rs.column_dimensions["A"].width = 52
rs.column_dimensions["B"].width = 30

wb.save(os.path.join(OUT, "findings.xlsx"))

# ── endpoint-inventory.xlsx (standalone, as specified) ───────────────────
wb2 = Workbook()
sheet(wb2, "Endpoint Inventory", eh, erows, [38, 13, 24, 15, 40, 34],
      wrap_cols=("Security Note",), first=True)
ws_e = wb2["Endpoint Inventory"]
for r in range(2, len(erows) + 2):
    if ws_e.cell(row=r, column=3).value == "NO":
        ws_e.cell(row=r, column=3).fill = warn
wb2.save(os.path.join(OUT, "endpoint-inventory.xlsx"))

print(f"findings           : {len(F)}  (C{counts['Critical']} H{counts['High']} "
      f"M{counts['Medium']} L{counts['Low']})")
print(f"endpoints          : {len(E)}  (unauthenticated flagged: {sum(1 for e in E if e[2]=='NO')})")
print(f"dependency issues  : {len(D)}")
print(f"security score     : {score}/100")
print(f"written            : {OUT}")
